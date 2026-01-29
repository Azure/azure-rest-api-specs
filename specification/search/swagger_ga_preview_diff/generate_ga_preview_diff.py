"""
GA vs Preview Diff Report Generator

Generates a single Excel workbook with two sheets comparing swagger versions:
1. Routes Diff - Operation-level differences
2. Models Diff - Schema/model-level differences

Comparisons (Nov preview is source of truth):
- GA (2025-09-01) vs Nov Preview (2025-11-01-preview)
- Aug Preview (2025-08-01-preview) vs Nov Preview (2025-11-01-preview)

Each sheet includes columns for both comparisons to help identify:
- What's new in Nov preview vs GA (for GA scoping)
- What's new in Nov preview vs Aug preview (to distinguish truly new from preview-only features)

Usage:
    python generate_ga_preview_diff.py
    python generate_ga_preview_diff.py --output-dir ./output
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(2)

from ref_resolver import RefResolver

# ============================================================================
# CONFIGURATION
# ============================================================================

# File paths - relative to this script
SCRIPT_DIR = Path(__file__).parent
SPEC_ROOT = SCRIPT_DIR.parent

# GA paths (2025-09-01)
GA_PATHS = {
    "searchservice": SPEC_ROOT
    / "data-plane/Search/stable/2025-09-01/searchservice.json",
    "searchindex": SPEC_ROOT / "data-plane/Search/stable/2025-09-01/searchindex.json",
}

# Aug Preview paths (2025-08-01-preview)
AUG_PREVIEW_PATHS = {
    "searchservice": SPEC_ROOT
    / "data-plane/Search/preview/2025-08-01-preview/searchservice.json",
    "searchindex": SPEC_ROOT
    / "data-plane/Search/preview/2025-08-01-preview/searchindex.json",
}

# Nov Preview paths (2025-11-01-preview)
PREVIEW_PATHS = {
    "searchservice": SPEC_ROOT
    / "data-plane/Search/preview/2025-11-01-preview_unmigrated/searchservice.json",
    "searchindex": SPEC_ROOT
    / "data-plane/Search/preview/2025-11-01-preview_unmigrated/searchindex.json",
    "knowledgebase": SPEC_ROOT
    / "data-plane/Search/preview/2025-11-01-preview_unmigrated/knowledgebase.json",
}


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================


def load_swagger_file(file_path: Path) -> Optional[Dict[str, Any]]:
    """Load and parse a swagger JSON file. Returns None if file doesn't exist."""
    if not file_path.exists():
        return None

    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_path(path: str) -> str:
    """Normalize API path for comparison."""
    return path.strip()


def get_operation_key(path: str, method: str) -> str:
    """Generate unique key for operation (path + method)."""
    return f"{normalize_path(path)}::{method.upper()}"


def resolve_parameter_ref(
    param: Dict[str, Any], swagger: Dict[str, Any]
) -> Dict[str, Any]:
    """Resolve parameter $ref to actual definition."""
    if "$ref" not in param:
        return param

    ref_path = param["$ref"]
    if not ref_path.startswith("#/"):
        return param

    parts = ref_path[2:].split("/")
    current = swagger
    for part in parts:
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            return param

    return current if isinstance(current, dict) else param


def get_param_signature(param: Dict[str, Any], swagger: Dict[str, Any]) -> str:
    """
    Get parameter signature for comparison.
    Format: {in}:{name}
    """
    resolved = resolve_parameter_ref(param, swagger)
    name = resolved.get("name", "")
    in_loc = resolved.get("in", "")
    return f"{in_loc}:{name}"


def format_param_display(param: Dict[str, Any], swagger: Dict[str, Any]) -> str:
    """
    Format parameter for display in diff.
    Format: {location}: {name} ({required/optional}) {type}
    """
    resolved = resolve_parameter_ref(param, swagger)

    name = resolved.get("name", "")
    in_loc = resolved.get("in", "")
    required = resolved.get("required", False)
    req_text = "required" if required else "optional"

    # Get type
    param_type = resolved.get("type", "")
    if "schema" in resolved:
        schema = resolved["schema"]
        if "$ref" in schema:
            param_type = schema["$ref"].split("/")[-1]
        else:
            param_type = schema.get("type", "object")

    return f"{in_loc}: {name} ({req_text}) {param_type}"


# ============================================================================
# OPERATION EXTRACTION
# ============================================================================


def extract_operations(swagger: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Extract all operations from swagger.

    Returns dict keyed by operation_key with:
    - path, method, operationId
    - parameters (list of param dicts)
    - responses (dict of status_code -> response)
    """
    operations = {}
    paths = swagger.get("paths", {})
    http_methods = {"get", "post", "put", "delete", "patch", "options", "head"}

    for path, path_obj in paths.items():
        path_params = path_obj.get("parameters", [])

        for method, operation in path_obj.items():
            method_lower = method.lower()
            if method_lower not in http_methods:
                continue

            # Combine path and operation parameters
            op_params = operation.get("parameters", [])
            all_params = path_params + op_params

            op_key = get_operation_key(path, method_lower)
            operations[op_key] = {
                "path": path,
                "method": method_lower.upper(),
                "operationId": operation.get("operationId", ""),
                "parameters": all_params,
                "responses": operation.get("responses", {}),
            }

    return operations


# ============================================================================
# ROUTES DIFF COMPUTATION
# ============================================================================


def compute_route_diff_generic(
    nov_op: Dict[str, Any],
    baseline_op: Optional[Dict[str, Any]],
    nov_swagger: Dict[str, Any],
    baseline_swagger: Optional[Dict[str, Any]],
    baseline_label: str,  # e.g., "2025-09-01" or "2025-08-01-preview"
) -> Tuple[str, str]:
    """
    Compute route-level differences between Nov preview and a baseline.

    Args:
        nov_op: Nov preview operation
        baseline_op: Baseline operation (GA or Aug preview)
        nov_swagger: Nov preview swagger
        baseline_swagger: Baseline swagger
        baseline_label: Label for presence string (e.g., "2025-09-01")

    Returns:
        - presence: "present in {baseline_label}" or "new in 2025-11-01-preview"
        - detail_diff: Description of route-level changes
    """
    if baseline_op is None:
        return f"new in 2025-11-01-preview", "new operation"

    presence = f"present in {baseline_label}"
    changes = []

    # Compare parameters
    nov_params = nov_op["parameters"]
    baseline_params = baseline_op["parameters"]

    nov_param_sigs = {get_param_signature(p, nov_swagger) for p in nov_params}
    baseline_param_sigs = {
        get_param_signature(p, baseline_swagger) for p in baseline_params
    }

    added_params = nov_param_sigs - baseline_param_sigs
    removed_params = baseline_param_sigs - nov_param_sigs

    if added_params or removed_params:
        param_changes = []
        if added_params:
            for sig in sorted(added_params):
                for p in nov_params:
                    if get_param_signature(p, nov_swagger) == sig:
                        param_changes.append(
                            f"+ {format_param_display(p, nov_swagger)}"
                        )
                        break
        if removed_params:
            for sig in sorted(removed_params):
                for p in baseline_params:
                    if get_param_signature(p, baseline_swagger) == sig:
                        param_changes.append(
                            f"- {format_param_display(p, baseline_swagger)}"
                        )
                        break

        if param_changes:
            changes.append("Parameters: " + "; ".join(param_changes))

    # Compare response status codes
    nov_responses = set(nov_op["responses"].keys())
    baseline_responses = set(baseline_op["responses"].keys())

    added_responses = nov_responses - baseline_responses
    removed_responses = baseline_responses - nov_responses

    if added_responses or removed_responses:
        resp_changes = []
        if added_responses:
            for code in sorted(added_responses):
                resp_changes.append(f"+ {code}")
        if removed_responses:
            for code in sorted(removed_responses):
                resp_changes.append(f"- {code}")

        if resp_changes:
            changes.append("Responses: " + "; ".join(resp_changes))

    # Build detail diff
    if not changes:
        if "2025-08" in baseline_label:
            detail_diff = "no changes from 2025-08-01-preview"
        else:
            detail_diff = "no changes for existing operation"
    else:
        detail_diff = " | ".join(changes)

    return presence, detail_diff


def compute_route_diff(
    preview_op: Dict[str, Any],
    ga_op: Optional[Dict[str, Any]],
    preview_swagger: Dict[str, Any],
    ga_swagger: Optional[Dict[str, Any]],
) -> Tuple[str, str]:
    """
    Compute route-level differences between preview and GA operations.
    (Wrapper for backward compatibility)
    """
    return compute_route_diff_generic(
        preview_op, ga_op, preview_swagger, ga_swagger, "2025-09-01"
    )


# ============================================================================
# MODEL EXTRACTION
# ============================================================================


def extract_models(swagger: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Extract all model definitions from swagger.

    Returns dict keyed by model name with definition schema.
    """
    return swagger.get("definitions", {})


def get_property_type(prop_schema: Dict[str, Any], swagger: Dict[str, Any]) -> str:
    """Get formatted type string for a property."""
    if "$ref" in prop_schema:
        return prop_schema["$ref"].split("/")[-1]

    prop_type = prop_schema.get("type", "")

    if prop_type == "array":
        items = prop_schema.get("items", {})
        items_type = get_property_type(items, swagger)
        return f"array of {items_type}"

    if prop_schema.get("format"):
        return f"{prop_type} (format: {prop_schema['format']})"

    return prop_type or "object"


# ============================================================================
# MODELS DIFF COMPUTATION
# ============================================================================


def compute_single_model_diff(
    model_name: str,
    prop_name: str,
    nov_model: Optional[Dict[str, Any]],
    baseline_model: Optional[Dict[str, Any]],
    nov_swagger: Dict[str, Any],
    baseline_swagger: Dict[str, Any],
    baseline_label: str,  # "GA" or "Aug"
) -> str:
    """
    Compute a concise diff description for Aug preview comparison.

    Returns: One-line description relative to the baseline (Aug or GA)
    """
    # Model-level diff
    if not prop_name:
        if nov_model and not baseline_model:
            return "new model"
        elif baseline_model and not nov_model:
            return "model removed in Nov"
        else:
            return (
                "no changes from 2025-08-01-preview"
                if baseline_label == "Aug"
                else "no model changes"
            )

    # Property-level diff
    if not nov_model or not baseline_model:
        return ""

    nov_props = nov_model.get("properties", {})
    baseline_props = baseline_model.get("properties", {})

    nov_required = set(nov_model.get("required", []))
    baseline_required = set(baseline_model.get("required", []))

    nov_prop = nov_props.get(prop_name)
    baseline_prop = baseline_props.get(prop_name)

    # New property
    if nov_prop and not baseline_prop:
        is_required = prop_name in nov_required
        return f"property added in Nov vs {baseline_label} ({'required' if is_required else 'optional'})"

    # Removed property
    if baseline_prop and not nov_prop:
        return f"property removed in Nov vs {baseline_label}"

    # Property exists in both - check changes
    if nov_prop and baseline_prop:
        # Type change
        nov_type = get_property_type(nov_prop, nov_swagger)
        baseline_type = get_property_type(baseline_prop, baseline_swagger)

        if nov_type != baseline_type:
            return (
                f"type changed in Nov vs {baseline_label}: {baseline_type} → {nov_type}"
            )

        # Required change
        was_required = prop_name in baseline_required
        is_required = prop_name in nov_required

        if was_required != is_required:
            if is_required:
                return f"optional → required in Nov vs {baseline_label}"
            else:
                return f"required → optional in Nov vs {baseline_label}"

    return f"no changes from 2025-08-01-preview" if baseline_label == "Aug" else ""


def compute_model_diffs_unified(
    nov_models: Dict[str, Dict[str, Any]],
    sep_models: Dict[str, Dict[str, Any]],
    aug_models: Dict[str, Dict[str, Any]],
    nov_swagger: Dict[str, Any],
    sep_swagger: Dict[str, Any],
    aug_swagger: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Compute unified model diffs with Nov preview as baseline.

    ROW INCLUSION RULE:
    - Rows are generated ONLY from Sep GA vs Nov preview differences
    - Aug preview columns are populated as annotations for existing rows

    Returns list of diff records with new structure:
    - model_name
    - property_name
    - nov_value (2025-11-01-preview)
    - sep_change_type (Change Type Sep vs Nov)
    - sep_value (2025-09-01 GA)
    - aug_change_type (Change Type Aug vs Nov) - annotation only
    - aug_value (2025-08-01-preview) - annotation only
    - sep_impact (Impact Sep vs Nov)
    """
    diffs = []

    # Only iterate over models that exist in Nov or Sep (not Aug-only models)
    all_model_names = set(nov_models.keys()) | set(sep_models.keys())

    for model_name in sorted(all_model_names):
        nov_model = nov_models.get(model_name)
        sep_model = sep_models.get(model_name)
        aug_model = aug_models.get(model_name)

        # Skip model-level rows entirely - only generate property-level rows
        # Don't process properties if model doesn't exist in Nov
        if not nov_model:
            continue

        # Property-level diffs (only if model exists in Nov)
        if nov_model:
            nov_props = nov_model.get("properties", {})
            nov_required = set(nov_model.get("required", []))

            sep_props = sep_model.get("properties", {}) if sep_model else {}
            sep_required = set(sep_model.get("required", [])) if sep_model else set()

            aug_props = aug_model.get("properties", {}) if aug_model else {}
            aug_required = set(aug_model.get("required", [])) if aug_model else set()

            # Only iterate over properties that exist in Nov or Sep (not Aug-only)
            all_props = set(nov_props.keys()) | set(sep_props.keys())

            for prop_name in sorted(all_props):
                nov_prop = nov_props.get(prop_name)
                sep_prop = sep_props.get(prop_name)
                aug_prop = aug_props.get(prop_name)

                # Determine Nov value
                if nov_prop:
                    nov_type = get_property_type(nov_prop, nov_swagger)
                    is_required = prop_name in nov_required
                    nov_value = (
                        f"{nov_type} ({'required' if is_required else 'optional'})"
                    )
                else:
                    nov_value = "(not present)"

                # Compare with Sep GA to determine if row should be included
                sep_change_type = ""
                sep_value = ""
                sep_impact = ""

                if nov_prop and not sep_prop:
                    # New property in Nov vs Sep
                    sep_change_type = "property added"
                    sep_value = "(not present)"
                    sep_impact = (
                        "breaking" if prop_name in nov_required else "non-breaking"
                    )

                elif sep_prop and not nov_prop:
                    # Property removed in Nov vs Sep
                    sep_change_type = "property removed"
                    sep_type = get_property_type(sep_prop, sep_swagger)
                    was_required = prop_name in sep_required
                    sep_value = (
                        f"{sep_type} ({'required' if was_required else 'optional'})"
                    )
                    sep_impact = "breaking"

                elif sep_prop and nov_prop:
                    # Both exist - check for changes
                    nov_type = get_property_type(nov_prop, nov_swagger)
                    sep_type = get_property_type(sep_prop, sep_swagger)

                    was_required_sep = prop_name in sep_required
                    is_required_nov = prop_name in nov_required

                    sep_value = (
                        f"{sep_type} ({'required' if was_required_sep else 'optional'})"
                    )

                    if nov_type != sep_type:
                        sep_change_type = "type changed"
                        sep_impact = "breaking"
                    elif was_required_sep != is_required_nov:
                        if is_required_nov:
                            sep_change_type = "optional → required"
                            sep_impact = "breaking"
                        else:
                            sep_change_type = "required → optional"
                            sep_impact = "non-breaking"
                    # else: no Sep vs Nov change, don't add row

                # ONLY add row if there's a Sep vs Nov change
                if not sep_change_type:
                    continue

                # Now compute Aug annotation for this row
                aug_change_type = ""
                aug_value = ""

                if nov_prop and not aug_prop:
                    aug_change_type = "property added"
                    aug_value = "(not present)"
                elif aug_prop and not nov_prop:
                    aug_change_type = "property removed"
                    aug_type = get_property_type(aug_prop, aug_swagger)
                    was_required_aug = prop_name in aug_required
                    aug_value = (
                        f"{aug_type} ({'required' if was_required_aug else 'optional'})"
                    )
                elif aug_prop and nov_prop:
                    # Both exist - check for changes
                    nov_type = get_property_type(nov_prop, nov_swagger)
                    aug_type = get_property_type(aug_prop, aug_swagger)

                    was_required_aug = prop_name in aug_required
                    is_required_nov = prop_name in nov_required

                    aug_value = (
                        f"{aug_type} ({'required' if was_required_aug else 'optional'})"
                    )

                    if nov_type != aug_type:
                        aug_change_type = "type changed"
                    elif was_required_aug != is_required_nov:
                        if is_required_nov:
                            aug_change_type = "optional → required"
                        else:
                            aug_change_type = "required → optional"
                    # else: no Aug vs Nov change, leave blank
                else:
                    # Property doesn't exist in Aug
                    aug_value = "(not present)"

                # Add the row (we know sep_change_type is not empty)
                diffs.append(
                    {
                        "model_name": model_name,
                        "property_name": prop_name,
                        "nov_value": nov_value,
                        "sep_change_type": sep_change_type,
                        "sep_value": sep_value,
                        "aug_change_type": aug_change_type,
                        "aug_value": aug_value,
                        "sep_impact": sep_impact,
                    }
                )

    return diffs


def compute_model_diffs(
    preview_models: Dict[str, Dict[str, Any]],
    ga_models: Dict[str, Dict[str, Any]],
    preview_swagger: Dict[str, Any],
    ga_swagger: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Compute model-level differences.

    Returns list of diff records with:
    - model_name: Name of the model
    - property_name: Name of the property (or "" for model-level changes)
    - change_type: added_model, removed_model, added_property, removed_property,
                   type_changed, required_changed, etc.
    - ga_value: Value in GA
    - preview_value: Value in Preview
    - impact: breaking, non-breaking, unknown
    """
    diffs = []

    all_model_names = set(preview_models.keys()) | set(ga_models.keys())

    for model_name in sorted(all_model_names):
        preview_model = preview_models.get(model_name)
        ga_model = ga_models.get(model_name)

        # New model in preview
        if preview_model and not ga_model:
            diffs.append(
                {
                    "model_name": model_name,
                    "property_name": "",
                    "change_type": "added_model",
                    "ga_value": "(not present)",
                    "preview_value": "new model",
                    "impact": "non-breaking",
                }
            )
            continue

        # Removed model in preview
        if ga_model and not preview_model:
            diffs.append(
                {
                    "model_name": model_name,
                    "property_name": "",
                    "change_type": "removed_model",
                    "ga_value": "model exists",
                    "preview_value": "(not present)",
                    "impact": "breaking",
                }
            )
            continue

        # Model exists in both - compare properties
        preview_props = preview_model.get("properties", {})
        ga_props = ga_model.get("properties", {})

        preview_required = set(preview_model.get("required", []))
        ga_required = set(ga_model.get("required", []))

        all_prop_names = set(preview_props.keys()) | set(ga_props.keys())

        for prop_name in sorted(all_prop_names):
            preview_prop = preview_props.get(prop_name)
            ga_prop = ga_props.get(prop_name)

            # New property
            if preview_prop and not ga_prop:
                prop_type = get_property_type(preview_prop, preview_swagger)
                is_required = prop_name in preview_required

                diffs.append(
                    {
                        "model_name": model_name,
                        "property_name": prop_name,
                        "change_type": "added_property",
                        "ga_value": "(not present)",
                        "preview_value": f"{prop_type} ({'required' if is_required else 'optional'})",
                        "impact": "breaking" if is_required else "non-breaking",
                    }
                )
                continue

            # Removed property
            if ga_prop and not preview_prop:
                prop_type = get_property_type(ga_prop, ga_swagger)
                is_required = prop_name in ga_required

                diffs.append(
                    {
                        "model_name": model_name,
                        "property_name": prop_name,
                        "change_type": "removed_property",
                        "ga_value": f"{prop_type} ({'required' if is_required else 'optional'})",
                        "preview_value": "(not present)",
                        "impact": "breaking",
                    }
                )
                continue

            # Property exists in both - check for changes
            preview_type = get_property_type(preview_prop, preview_swagger)
            ga_type = get_property_type(ga_prop, ga_swagger)

            # Type changed
            if preview_type != ga_type:
                diffs.append(
                    {
                        "model_name": model_name,
                        "property_name": prop_name,
                        "change_type": "type_changed",
                        "ga_value": ga_type,
                        "preview_value": preview_type,
                        "impact": "breaking",
                    }
                )

            # Required status changed
            was_required = prop_name in ga_required
            is_required = prop_name in preview_required

            if was_required != is_required:
                impact = "breaking" if is_required else "non-breaking"
                diffs.append(
                    {
                        "model_name": model_name,
                        "property_name": prop_name,
                        "change_type": "required_changed",
                        "ga_value": "required" if was_required else "optional",
                        "preview_value": "required" if is_required else "optional",
                        "impact": impact,
                    }
                )

    return diffs


# ============================================================================
# EXCEL GENERATION
# ============================================================================

# Color definitions for row highlighting
COLORS = {
    "light_blue": "D6EAF8",  # New operations
    "light_yellow": "FFF9C4",  # Route changes or SDK-impacting model changes
    "light_red": "FFCDD2",  # Breaking model changes
    "light_gray": "EEEEEE",  # Non-impacting changes
}


def create_routes_sheet(
    wb: Workbook, spec_name: str, route_diffs: List[Dict[str, Any]]
):
    """Create Routes Diff sheet."""
    if "Routes Diff" in wb.sheetnames:
        ws = wb["Routes Diff"]
    else:
        ws = wb.create_sheet("Routes Diff")

    # Headers
    headers = [
        "Spec",
        "2025-11-01-preview Path",
        "Method",
        "Operation ID",
        "2025-09-01 GA Presence",
        "Detail Diff",
        "2025-08-01-preview Presence",
        "Detailed Diff from 2025-08-01-preview",
        "2026-04-01 GA Decision",
    ]

    # Write headers (only once if sheet is new)
    if ws.max_row == 1:
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Set column widths
        ws.column_dimensions["A"].width = 15  # Spec
        ws.column_dimensions["B"].width = 40  # Path
        ws.column_dimensions["C"].width = 10  # Method
        ws.column_dimensions["D"].width = 35  # Operation ID
        ws.column_dimensions["E"].width = 25  # GA Presence
        ws.column_dimensions["F"].width = 60  # Detail Diff
        ws.column_dimensions["G"].width = 25  # Aug Preview Presence
        ws.column_dimensions["H"].width = 60  # Aug Preview Diff
        ws.column_dimensions["I"].width = 25  # Decision

    # Find next row
    next_row = ws.max_row + 1

    # Write data
    for diff in route_diffs:
        ws.cell(row=next_row, column=1, value=spec_name)
        ws.cell(row=next_row, column=2, value=diff["path"])
        ws.cell(row=next_row, column=3, value=diff["method"])
        ws.cell(row=next_row, column=4, value=diff["operationId"])
        ws.cell(row=next_row, column=5, value=diff["presence"])

        # Detail Diff with wrap text
        detail_cell = ws.cell(row=next_row, column=6, value=diff["detail_diff"])
        detail_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Aug preview columns
        ws.cell(row=next_row, column=7, value=diff.get("aug_presence", ""))
        aug_diff_cell = ws.cell(row=next_row, column=8, value=diff.get("aug_diff", ""))
        aug_diff_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=next_row, column=9, value="")  # Decision - blank

        # Apply row coloring based on diff type
        row_color = None
        presence = diff["presence"]
        detail_diff = diff["detail_diff"]

        # 1. New operation (light blue)
        if presence == "new in 2025-11-01-preview" and detail_diff == "new operation":
            row_color = COLORS["light_blue"]

        # 2. Existing operation with route-level changes (light yellow)
        elif (
            presence == "present in 2025-09-01"
            and detail_diff != "no changes for existing operation"
        ):
            row_color = COLORS["light_yellow"]

        # 3. No changes - keep white (no color)
        # else: row_color = None (default)

        # Apply color to entire row
        if row_color:
            fill = PatternFill(
                start_color=row_color, end_color=row_color, fill_type="solid"
            )
            for col in range(1, 10):  # Columns A-I
                ws.cell(row=next_row, column=col).fill = fill

        next_row += 1


def create_models_sheet(
    wb: Workbook, spec_name: str, model_diffs: List[Dict[str, Any]]
):
    """Create Models Diff sheet."""
    if "Models Diff" in wb.sheetnames:
        ws = wb["Models Diff"]
    else:
        ws = wb.create_sheet("Models Diff")

    # Headers
    headers = [
        "Spec",
        "Model Name",
        "Property Name",
        "2025-11-01-preview Value",
        "Change Type (Sep vs Nov)",
        "2025-09-01 GA Value",
        "Change Type (Aug vs Nov)",
        "2025-08-01-preview Value",
        "Impact (Sep vs Nov)",
        "2026-04-01 GA Decision",
    ]

    # Write headers (only once if sheet is new)
    if ws.max_row == 1:
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Set column widths
        ws.column_dimensions["A"].width = 15  # Spec
        ws.column_dimensions["B"].width = 30  # Model Name
        ws.column_dimensions["C"].width = 25  # Property Name
        ws.column_dimensions["D"].width = 35  # Nov Value
        ws.column_dimensions["E"].width = 25  # Change Type (Sep vs Nov)
        ws.column_dimensions["F"].width = 35  # Sep Value
        ws.column_dimensions["G"].width = 25  # Change Type (Aug vs Nov)
        ws.column_dimensions["H"].width = 35  # Aug Value
        ws.column_dimensions["I"].width = 18  # Impact (Sep vs Nov)
        ws.column_dimensions["J"].width = 25  # Decision

    # Find next row
    next_row = ws.max_row + 1

    # Write data
    for diff in model_diffs:
        ws.cell(row=next_row, column=1, value=spec_name)
        ws.cell(row=next_row, column=2, value=diff["model_name"])
        ws.cell(row=next_row, column=3, value=diff["property_name"])
        ws.cell(row=next_row, column=4, value=diff["nov_value"])
        ws.cell(row=next_row, column=5, value=diff["sep_change_type"])
        ws.cell(row=next_row, column=6, value=diff["sep_value"])
        ws.cell(row=next_row, column=7, value=diff["aug_change_type"])
        ws.cell(row=next_row, column=8, value=diff["aug_value"])
        ws.cell(row=next_row, column=9, value=diff["sep_impact"])
        ws.cell(row=next_row, column=10, value="")  # Decision - blank

        # Apply row coloring based on Sep change type and impact
        row_color = None
        sep_change_type = diff.get("sep_change_type", "")
        sep_impact = diff.get("sep_impact", "")

        # 1. Potentially breaking model changes (light red)
        breaking_types = ["property removed", "type changed"]
        if sep_change_type in breaking_types or sep_impact == "breaking":
            row_color = COLORS["light_red"]

        # 2. Compatibility-sensitive changes (light yellow)
        elif sep_change_type == "property added" or "→" in sep_change_type:
            # Check if it's breaking (required changes)
            if sep_impact == "breaking":
                row_color = COLORS["light_red"]
            elif sep_impact == "non-breaking":
                row_color = COLORS["light_yellow"]

        # 3. Non-breaking additions at model level (light gray for informational)
        elif (
            diff["property_name"] == "(model-level)"
            and sep_change_type == "property added"
        ):
            row_color = COLORS["light_gray"]

        # Apply color to entire row
        if row_color:
            fill = PatternFill(
                start_color=row_color, end_color=row_color, fill_type="solid"
            )
            for col in range(1, 11):  # Columns A-J
                ws.cell(row=next_row, column=col).fill = fill

        next_row += 1


# ============================================================================
# MAIN
# ============================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Generate GA vs Preview diff Excel report with Routes and Models sheets"
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Output directory for Excel file (default: output)",
    )
    args = parser.parse_args()

    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Process each spec
    specs_to_process = ["searchservice", "searchindex", "knowledgebase"]

    for spec_name in specs_to_process:
        print(f"\nProcessing {spec_name}...")

        # Load preview (always exists)
        preview_path = PREVIEW_PATHS[spec_name]
        preview_swagger = load_swagger_file(preview_path)

        if not preview_swagger:
            print(f"  Warning: Preview file not found: {preview_path}")
            continue

        # Load GA (may not exist for knowledgebase)
        ga_swagger = None
        if spec_name in GA_PATHS:
            ga_path = GA_PATHS[spec_name]
            ga_swagger = load_swagger_file(ga_path)
            if not ga_swagger:
                print(f"  Warning: GA file not found: {ga_path}")

        # Load Aug Preview (may not exist for knowledgebase)
        aug_swagger = None
        if spec_name in AUG_PREVIEW_PATHS:
            aug_path = AUG_PREVIEW_PATHS[spec_name]
            aug_swagger = load_swagger_file(aug_path)
            if not aug_swagger:
                print(f"  Warning: Aug Preview file not found: {aug_path}")

        # Extract operations
        preview_ops = extract_operations(preview_swagger)
        ga_ops = extract_operations(ga_swagger) if ga_swagger else {}
        aug_ops = extract_operations(aug_swagger) if aug_swagger else {}

        print(f"  Preview operations: {len(preview_ops)}")
        print(f"  GA operations: {len(ga_ops)}")
        print(f"  Aug Preview operations: {len(aug_ops)}")

        # Compute route diffs
        route_diffs = []
        for op_key, preview_op in sorted(preview_ops.items()):
            ga_op = ga_ops.get(op_key)
            presence, detail_diff = compute_route_diff(
                preview_op, ga_op, preview_swagger, ga_swagger
            )

            # Also compute Aug preview diff
            aug_op = aug_ops.get(op_key)
            aug_presence, aug_diff = compute_route_diff_generic(
                preview_op, aug_op, preview_swagger, aug_swagger, "2025-08-01-preview"
            )

            route_diffs.append(
                {
                    "path": preview_op["path"],
                    "method": preview_op["method"],
                    "operationId": preview_op["operationId"],
                    "presence": presence,
                    "detail_diff": detail_diff,
                    "aug_presence": aug_presence,
                    "aug_diff": aug_diff,
                }
            )

        print(f"  Route diffs: {len(route_diffs)}")

        # Extract models
        preview_models = extract_models(preview_swagger)
        ga_models = extract_models(ga_swagger) if ga_swagger else {}
        aug_models = extract_models(aug_swagger) if aug_swagger else {}

        print(f"  Preview models: {len(preview_models)}")
        print(f"  GA models: {len(ga_models)}")
        print(f"  Aug Preview models: {len(aug_models)}")

        # Compute model diffs (Sep GA vs Nov Preview only)
        model_diffs = compute_model_diffs(
            preview_models,
            ga_models,
            preview_swagger,
            ga_swagger if ga_swagger else {},
        )

        print(f"  Model diffs: {len(model_diffs)}")

        # Annotate Sep vs Nov diffs with Aug preview information
        for diff in model_diffs:
            model_name = diff["model_name"]
            prop_name = diff["property_name"]
            
            nov_model = preview_models.get(model_name)
            aug_model = aug_models.get(model_name)
            
            # Convert old format to new format
            nov_value = diff["preview_value"]
            sep_change_type = diff["change_type"]
            sep_value = diff["ga_value"]
            sep_impact = diff["impact"]
            
            # Determine Aug annotation
            aug_change_type = ""
            aug_value = ""
            
            if not aug_model:
                # Model doesn't exist in Aug
                if nov_model:
                    aug_change_type = "property added"
                    aug_value = "(not present)"
            elif not nov_model:
                # Model removed in Nov (exists in Aug)
                aug_change_type = "property removed"
                aug_value = "model exists"
            elif prop_name == "":
                # Model-level diff - model exists in both Aug and Nov
                aug_value = "model exists"
                aug_change_type = ""
            else:
                # Property-level - check Aug vs Nov
                aug_props = aug_model.get("properties", {})
                nov_props = nov_model.get("properties", {})
                
                aug_required = set(aug_model.get("required", []))
                nov_required = set(nov_model.get("required", []))
                
                aug_prop = aug_props.get(prop_name)
                nov_prop = nov_props.get(prop_name)
                
                if aug_prop and not nov_prop:
                    # Property removed in Nov vs Aug
                    aug_type = get_property_type(aug_prop, aug_swagger)
                    was_required = prop_name in aug_required
                    aug_value = f"{aug_type} ({'required' if was_required else 'optional'})"
                    aug_change_type = "property removed"
                elif not aug_prop and nov_prop:
                    # Property added in Nov vs Aug
                    aug_value = "(not present)"
                    aug_change_type = "property added"
                elif aug_prop and nov_prop:
                    # Both exist - check for changes
                    aug_type = get_property_type(aug_prop, aug_swagger)
                    nov_type = get_property_type(nov_prop, preview_swagger)
                    
                    was_required_aug = prop_name in aug_required
                    is_required_nov = prop_name in nov_required
                    
                    aug_value = f"{aug_type} ({'required' if was_required_aug else 'optional'})"
                    
                    if aug_type != nov_type:
                        aug_change_type = "type changed"
                    elif was_required_aug != is_required_nov:
                        if is_required_nov:
                            aug_change_type = "optional → required"
                        else:
                            aug_change_type = "required → optional"
                    # else: no change, leave blank
            
            # Update diff with new structure
            diff["nov_value"] = nov_value
            diff["sep_change_type"] = sep_change_type
            diff["sep_value"] = sep_value
            diff["aug_change_type"] = aug_change_type
            diff["aug_value"] = aug_value
            diff["sep_impact"] = sep_impact

        # Add to sheets
        if route_diffs:
            create_routes_sheet(wb, spec_name, route_diffs)

        if model_diffs:
            create_models_sheet(wb, spec_name, model_diffs)

    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"ga_preview_diff_{timestamp}.xlsx"

    # Ensure we have at least one sheet
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    wb.save(output_file)
    print(f"\n✓ Excel report saved: {output_file}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
