"""
GA vs Preview Swagger Discrepancy Checker

This CLI tool compares GA (stable) swagger files against Preview swagger files
to identify breaking changes and discrepancies. The comparison logic mirrors
the semantic equivalency checker in ../swagger_equiv/semantic_equiv but is
fully self-contained without shared dependencies.

Key comparison aspects (mirroring semantic_equiv):
1. Canonicalization - removes doc-only fields (description, summary, examples, tags)
2. Normalization - sorts/deduplicates set-like arrays (consumes, produces, schemes)
3. Deep comparison - paths, operations, parameters, schemas, constraints
4. Severity classification - breaking vs non-breaking vs unknown changes
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(2)

# Import comparison modules
from canonicalize import canonicalize_swagger
from compare import ComparisonResult, compare_swagger_files
from severity import SeverityLevel, classify_severity


def load_swagger_file(file_path: str) -> Dict[str, Any]:
    """Load and parse a swagger JSON file."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Swagger file not found: {file_path}")

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def categorize_difference(diff: Dict[str, Any]) -> str:
    """Categorize a difference into operation/parameter/response/definition/schema/other."""
    diff_type = diff["type"]
    context = diff.get("context", "")

    # Check context first for definition
    if "definition:" in context and not ("[" in context and "]" in context):
        # This is a definition-level change (not within an operation)
        return "definition"
    elif any(x in diff_type for x in ["path", "method", "operation"]):
        return "operation"
    elif "parameter" in diff_type:
        return "parameter"
    elif "response" in diff_type:
        return "response"
    elif any(
        x in diff_type for x in ["schema", "definition", "properties", "required"]
    ):
        return "schema"
    else:
        return "other"


def extract_identifier(diff: Dict[str, Any]) -> str:
    """
    Extract detailed identifier from the difference.

    Format includes: path, method, operation ID, parameter/definition name
    Examples:
    - "/docs [get] operationId:Documents_SearchGet param:queryLanguage"
    - "definition:SearchRequest property:hybridSearch"
    - "/indexes/{indexName} [put] operationId:Indexes_CreateOrUpdate"
    """
    context = diff.get("context", "")
    diff_type = diff["type"]
    message = diff.get("message", "")

    if not context:
        return diff_type

    # Build detailed identifier
    identifier_parts = []

    # Extract path and method if present
    if "[" in context and "]" in context:
        # Format: "/path [METHOD]" or "/path [METHOD] ..."
        path_method = context.split("]")[0] + "]"
        identifier_parts.append(path_method)

        # Extract operation ID from message if it contains "Operation ID"
        if "Operation ID" in message:
            # Pattern: "Operation ID changed: old -> new"
            op_parts = message.split("Operation ID")
            if len(op_parts) > 1:
                op_info = op_parts[1].strip()
                if ":" in op_info:
                    op_id = op_info.split(":")[1].split("->")[0].strip()
                    identifier_parts.append(f"operationId:{op_id}")

        # Extract parameter or response info from context
        remaining = context.split("]", 1)[1] if "]" in context else ""
        if "param:" in remaining:
            param_name = remaining.split("param:")[1].split()[0]
            identifier_parts.append(f"param:{param_name}")
        elif "response:" in remaining:
            resp_code = remaining.split("response:")[1].split()[0]
            identifier_parts.append(f"response:{resp_code}")
        elif "request_body" in remaining:
            identifier_parts.append("request_body")

    # Handle definition/schema contexts
    elif "definition:" in context:
        # Format: "definition:ModelName" or "definition:ModelName property:propName"
        def_info = context.split("definition:")[1]

        # Split by spaces to get definition name
        def_parts = def_info.split()
        if def_parts:
            # Just use the definition name without the "definition:" prefix
            identifier_parts.append(def_parts[0])

            # Check for property, items, etc.
            if "property:" in def_info:
                prop_name = def_info.split("property:")[1].split()[0]
                identifier_parts.append(f"property:{prop_name}")
            elif "items" in def_info:
                identifier_parts.append("items")

    # If no specific format matched, use context as-is
    if not identifier_parts:
        identifier_parts.append(context)

    return " ".join(identifier_parts)


def extract_values(diff: Dict[str, Any]) -> Tuple[str, str]:
    """
    Extract GA and Preview values from the difference message.

    Returns (GA_value, Preview_value) tuple.
    """
    message = diff["message"]
    diff_type = diff["type"]

    # Pattern 1: "X changed: value1 -> value2"
    if " -> " in message and "changed" in message.lower():
        parts = message.split(" -> ")
        if len(parts) == 2:
            # Extract the part after the colon for GA value
            ga_val = parts[0].split(":")[-1].strip()
            preview_val = parts[1].strip()
            return ga_val, preview_val

    # Pattern 2: "missing in Preview: value"
    if "missing in" in message.lower() and "preview" in message.lower():
        # Present in GA, absent in Preview
        value = message.split(":")[-1].strip() if ":" in message else ""
        # Clean up the value
        value = value.replace(f"missing in Preview:", "").strip()
        return value, "(not present)"

    # Pattern 3: "extra in Preview: value"
    if "extra" in message.lower() and "preview" in message.lower():
        # Absent in GA, present in Preview
        # For parameters: "Extra parameter in Preview: paramName (in: query, required: False)"
        # We want to extract just "paramName"
        if "Extra parameter in Preview:" in message:
            # Split by the first colon to get the part after "Preview:"
            parts = message.split("Extra parameter in Preview:", 1)
            if len(parts) == 2:
                param_info = parts[1].strip()
                # Extract just the parameter name before the parenthesis
                if "(" in param_info:
                    value = param_info.split("(")[0].strip()
                else:
                    value = param_info
                return "(not present)", value

        # For other "extra" messages
        value = message.split(":")[-1].strip() if ":" in message else ""
        # Clean up the value - extract details if present
        if "(in:" in value or "(" in value:
            # Extract the part before parenthesis
            value = value.split("(")[0].strip()
        return "(not present)", value

    # Pattern 4: Properties/Required fields added/removed
    if "added:" in message.lower():
        # Extract the list of added items
        added_items = message.split("added:")[1].strip()
        return "(not present)", added_items

    if "removed:" in message.lower():
        # Extract the list of removed items
        removed_items = message.split("removed:")[1].strip()
        return removed_items, "(not present)"

    # Pattern 5: Comparison messages like "GA X vs Preview Y"
    if " vs " in message:
        parts = message.split(" vs ")
        if len(parts) == 2:
            ga_part = parts[0]
            preview_part = parts[1]

            # Extract values after "GA" or source name
            ga_val = ga_part.split()[-1] if ga_part.split() else ""
            preview_val = preview_part.split()[0] if preview_part.split() else ""

            return ga_val, preview_val

    # Default: return empty GA, full message as Preview (for additions)
    return "", message


def determine_disposition(diff: Dict[str, Any]) -> str:
    """
    Determine the disposition action for a discrepancy.

    Returns one of:
    - "Tooling false positive"
    - "Fix in TSP"
    - "GA exclude (preview-only)"
    - "GA include candidate"
    - "Needs PM decision"

    Args:
        diff: Difference dictionary with keys: category, message, severity, context, type

    Returns:
        Disposition string
    """
    category = categorize_difference(diff)
    message = diff.get("message", "").lower()
    summary = diff.get("message", "")
    severity = diff.get("severity", "")
    diff_type = diff.get("type", "")
    identifier = extract_identifier(diff)
    ga_val, preview_val = extract_values(diff)

    # Rule 1: Tooling false positive
    # Detect refactor-only changes like inline <-> $ref conversions
    if any(
        x in message
        for x in [
            "type changed: object -> none",
            "$ref",
            "schema changed from inline",
            "schema changed from reference",
        ]
    ):
        return "Tooling false positive"

    # Rule 2: Fix in TSP
    # Breaking changes that need to be fixed
    if severity == "breaking":
        if any(
            x in message
            for x in [
                "required fields added",
                "required fields changed",
                "required fields removed",
                "type changed",
                "properties removed",
            ]
        ):
            return "Fix in TSP"

    # Rule 3: GA exclude (preview-only)
    # Preview-only operations or definitions
    if category == "operation":
        if ga_val == "(not present)" or "extra path in preview" in message:
            return "GA exclude (preview-only)"

    if category in ("definition", "schema"):
        if "extra definition in preview" in message:
            return "GA exclude (preview-only)"

    # Rule 4: GA include candidate
    # Non-breaking additions to existing GA models
    if severity == "non-breaking" and "properties added" in message:
        # Check if this is extending an existing GA model
        # Heuristic: GA Value is present (not "(not present)")
        if ga_val and ga_val != "(not present)":
            return "GA include candidate"

        # Check if identifier references a core GA type
        core_ga_types = [
            "SearchIndex",
            "SearchField",
            "SearchIndexer",
            "SearchIndexerDataSource",
            "SearchIndexerSkillset",
            "ServiceLimits",
            "ServiceStatistics",
            "SynonymMap",
            "DocumentDebugInfo",
            "FacetResult",
            "SearchRequest",
            "SearchResult",
            "SearchDocumentsResult",
            "VectorQuery",
        ]

        if any(core_type in identifier for core_type in core_ga_types):
            return "GA include candidate"

    # Rule 5: Default fallback
    return "Needs PM decision"


def save_excel_report(results: Dict[str, ComparisonResult], output_dir: str) -> None:
    """Save comparison results as Excel file with separate sheets."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Add timestamp to filename to avoid overwriting
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = output_path / f"ga_preview_discrepancy_{timestamp}.xlsx"

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Define column headers
    headers = [
        "Category",
        "Identifier",
        "2025-09-01 GA Value",
        "2025-11-01-preview Value",
        "Difference Summary",
        "Severity",
        "2026-04-01 TSP Action",
        "Notes",
    ]

    # Severity colors
    severity_colors = {
        "breaking": "FFCCCC",  # Light red
        "non-breaking": "CCFFCC",  # Light green
        "unknown": "FFFFCC",  # Light yellow
    }

    for file_name, result in sorted(results.items()):
        # Create sheet name (remove .json extension)
        sheet_name = file_name.replace(".json", "")
        ws = wb.create_sheet(title=sheet_name)

        # Write headers with formatting
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows - sorted for deterministic output
        sorted_diffs = sorted(
            result.differences,
            key=lambda x: (
                x["severity"],  # breaking first, then non-breaking, then unknown
                categorize_difference(x),
                extract_identifier(x),
                x["type"],
            ),
        )

        for row_num, diff in enumerate(sorted_diffs, 2):
            category = categorize_difference(diff)
            identifier = extract_identifier(diff)
            ga_val, preview_val = extract_values(diff)
            summary = diff["message"]
            severity = diff["severity"]
            disposition = determine_disposition(diff)

            # Write row data
            ws.cell(row=row_num, column=1, value=category)
            ws.cell(row=row_num, column=2, value=identifier)
            ws.cell(row=row_num, column=3, value=ga_val)
            ws.cell(row=row_num, column=4, value=preview_val)
            ws.cell(row=row_num, column=5, value=summary)
            ws.cell(row=row_num, column=6, value=severity)
            ws.cell(row=row_num, column=7, value=disposition)
            ws.cell(row=row_num, column=8, value="")  # Notes column for reviewer

            # Apply severity color to entire row
            if severity in severity_colors:
                fill = PatternFill(
                    start_color=severity_colors[severity],
                    end_color=severity_colors[severity],
                    fill_type="solid",
                )
                for col_num in range(1, 9):
                    ws.cell(row=row_num, column=col_num).fill = fill

        # Auto-adjust column widths
        for col_num in range(1, 9):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

    # Save workbook
    wb.save(report_file)
    print(f"\n[OK] Excel report saved to: {report_file}")


def print_summary(results: Dict[str, ComparisonResult], verbose: bool = False) -> int:
    """
    Print comparison summary to console.

    Returns:
        Exit code (0 = no breaking changes, 1 = breaking changes found)
    """
    print("\n" + "=" * 60)
    print("=== GA vs Preview Comparison Report ===")
    print("=" * 60)

    has_breaking_changes = False
    total_breaking = 0
    total_non_breaking = 0
    total_unknown = 0

    for file_name, result in sorted(results.items()):
        print(f"\nFile: {file_name}")

        # Count by severity
        breaking = sum(1 for d in result.differences if d["severity"] == "breaking")
        non_breaking = sum(
            1 for d in result.differences if d["severity"] == "non-breaking"
        )
        unknown = sum(1 for d in result.differences if d["severity"] == "unknown")

        total_breaking += breaking
        total_non_breaking += non_breaking
        total_unknown += unknown

        print(f"  Breaking Changes: {breaking}")
        print(f"  Non-Breaking Changes: {non_breaking}")
        print(f"  Unknown: {unknown}")

        if breaking > 0:
            has_breaking_changes = True

        if verbose and result.differences:
            print(f"\n  Detailed Differences:")
            for diff in result.differences:
                severity_symbol = (
                    "[BREAK]"
                    if diff["severity"] == "breaking"
                    else "[WARN]"
                    if diff["severity"] == "unknown"
                    else "[INFO]"
                )
                print(f"    {severity_symbol} [{diff['type']}] {diff['message']}")
                if diff.get("context"):
                    print(f"       Context: {diff['context']}")

    print("\n" + "=" * 60)
    print(
        f"Total: {total_breaking} breaking, {total_non_breaking} non-breaking, {total_unknown} unknown"
    )
    print("=" * 60)

    if has_breaking_changes:
        print("\n[WARNING] BREAKING CHANGES DETECTED - Manual review required")
        return 1
    elif total_unknown > 0:
        print("\n[WARNING] Unknown changes detected - Manual review recommended")
        return 0
    else:
        print("\n[OK] No breaking changes detected")
        return 0


def main():
    """Main CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Compare GA and Preview swagger files for discrepancies"
    )

    # Get the script directory to build absolute paths
    script_dir = Path(__file__).parent
    spec_root = script_dir.parent.parent  # Go up to specification/ root

    # File selection
    parser.add_argument(
        "--file",
        choices=["searchindex", "searchservice", "both"],
        default="both",
        help="Which file pair to compare (default: both)",
    )

    # Custom paths for searchindex
    parser.add_argument(
        "--ga-index",
        default=str(
            spec_root / "search/data-plane/Search/stable/2025-09-01/searchindex.json"
        ),
        help="Path to GA searchindex.json",
    )
    parser.add_argument(
        "--preview-index",
        default=str(
            spec_root
            / "search/data-plane/Search/preview/2025-11-01-preview_unmigrated/searchindex.json"
        ),
        help="Path to Preview searchindex.json",
    )

    # Custom paths for searchservice
    parser.add_argument(
        "--ga-service",
        default=str(
            spec_root / "search/data-plane/Search/stable/2025-09-01/searchservice.json"
        ),
        help="Path to GA searchservice.json",
    )
    parser.add_argument(
        "--preview-service",
        default=str(
            spec_root
            / "search/data-plane/Search/preview/2025-11-01-preview_unmigrated/searchservice.json"
        ),
        help="Path to Preview searchservice.json",
    )

    # Output options
    parser.add_argument(
        "--output-dir",
        default="./output",
        help="Directory for output reports (default: ./output)",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Show detailed differences in console output",
    )

    args = parser.parse_args()

    # Determine which files to compare
    file_pairs = []
    if args.file in ["searchindex", "both"]:
        file_pairs.append(
            {
                "name": "searchindex.json",
                "ga_path": args.ga_index,
                "preview_path": args.preview_index,
            }
        )
    if args.file in ["searchservice", "both"]:
        file_pairs.append(
            {
                "name": "searchservice.json",
                "ga_path": args.ga_service,
                "preview_path": args.preview_service,
            }
        )

    try:
        all_results = {}

        for pair in file_pairs:
            print(f"\nComparing {pair['name']}...")
            print(f"  GA: {pair['ga_path']}")
            print(f"  Preview: {pair['preview_path']}")

            # Load files
            ga_swagger = load_swagger_file(pair["ga_path"])
            preview_swagger = load_swagger_file(pair["preview_path"])

            # Canonicalize both
            ga_canonical = canonicalize_swagger(ga_swagger)
            preview_canonical = canonicalize_swagger(preview_swagger)

            # Compare
            result = compare_swagger_files(
                ga_canonical, preview_canonical, ga_name="GA", preview_name="Preview"
            )

            all_results[pair["name"]] = result

        # Generate Excel report
        save_excel_report(all_results, args.output_dir)

        # Print summary and determine exit code
        exit_code = print_summary(all_results, verbose=args.verbose)

        sys.exit(exit_code)

    except FileNotFoundError as e:
        print(f"\n[ERROR] {e}", file=sys.stderr)
        sys.exit(2)
    except Exception as e:
        print(f"\n[ERROR] Unexpected error: {e}", file=sys.stderr)
        import traceback

        traceback.print_exc()
        sys.exit(2)


if __name__ == "__main__":
    main()
