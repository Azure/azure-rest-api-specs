"""
Operation Surface Report Generator

This tool generates an operation-centric API surface Excel report from Preview swagger files,
with diff annotations comparing against GA swagger files. The report focuses on operations
as the primary unit rather than schema-level differences.

Purpose:
- Use Preview (2025-11-01-preview) as the source of truth/baseline
- List all operations with detailed parameter and response information
- Annotate each operation with differences vs GA (2025-09-01) as guidance for 2026-04-01 GA decisions

Output:
- Single Excel workbook with 3 sheets: searchservice, searchindex, knowledgebase
- Each sheet has one row per operation (sorted by path then method)
- Columns: path, method, operationId, parameters (bullet list), responses (bullet list),
  GA presence ("present in 2025-09-01" or "new in 2025-11-01-preview"), detail diff

Usage:
    # Generate report (always creates 3 sheets in one workbook)
    python generate_operation_surface.py

    # Specify output directory
    python generate_operation_surface.py --output-dir ./reports

    # Custom paths (rarely needed)
    python generate_operation_surface.py \\
        --preview-index /path/to/preview/searchindex.json \\
        --ga-index /path/to/ga/searchindex.json
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(2)

from ref_resolver import RefResolver, normalize_schema

# ===== Helper Functions =====


def load_swagger_file(file_path: str) -> Dict[str, Any]:
    """Load and parse a swagger JSON file."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Swagger file not found: {file_path}")

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_path(path: str) -> str:
    """
    Normalize path template for consistent comparison.
    Keep path as-is since both specs should use the same style.
    """
    return path.strip()


def get_operation_key(path: str, method: str) -> str:
    """
    Generate a unique key for an operation based on normalized path + method.
    This is the primary key for matching operations between GA and Preview.
    """
    return f"{normalize_path(path)}::{method.upper()}"


def get_param_key(param: Dict[str, Any]) -> str:
    """
    Generate a unique key for a parameter based on name and location (in).
    """
    name = param.get("name", "")
    in_location = param.get("in", "")
    return f"{name}::{in_location}"


def format_type_info(
    schema: Dict[str, Any], resolver: Optional[RefResolver] = None
) -> str:
    """
    Format schema type information for display.
    Returns: type string, format string, or $ref name.

    Examples:
    - "string"
    - "integer (format: int32)"
    - "SearchDocumentsResult" (for $ref)
    - "array of string"
    """
    if not schema:
        return ""

    # Handle $ref
    if "$ref" in schema:
        ref_path = schema["$ref"]
        # Extract definition name from $ref path
        # Patterns: "#/definitions/ModelName" or just "ModelName"
        if "/" in ref_path:
            return ref_path.split("/")[-1]
        return ref_path

    # Resolve schema if resolver provided
    if resolver:
        schema = resolver.resolve_ref(schema) or schema

    # Get type
    schema_type = schema.get("type", "")
    schema_format = schema.get("format", "")

    # Handle array type
    if schema_type == "array":
        items = schema.get("items", {})
        items_type = format_type_info(items, resolver)
        return f"array of {items_type}"

    # Build type string
    if schema_format:
        return f"{schema_type} (format: {schema_format})"
    elif schema_type:
        return schema_type

    # Check for enum
    if "enum" in schema:
        enum_vals = schema["enum"]
        if len(enum_vals) <= 3:
            return f"enum: {', '.join(str(v) for v in enum_vals)}"
        else:
            return f"enum ({len(enum_vals)} values)"

    return "object" if not schema_type else schema_type


def resolve_parameter_ref(
    param: Dict[str, Any], swagger: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Resolve a parameter $ref to the actual parameter definition.

    Args:
        param: Parameter object (may contain $ref)
        swagger: Full swagger document

    Returns:
        Resolved parameter dictionary, or original if not a ref or not found
    """
    if "$ref" not in param:
        return param

    ref_path = param["$ref"]

    # Parse ref path (e.g., "#/parameters/ClientRequestIdParameter")
    if not ref_path.startswith("#/"):
        # Can't resolve external refs, return original with ref name
        return param

    # Split path and navigate through swagger document
    parts = ref_path[2:].split("/")  # Skip "#/" prefix

    current = swagger
    for part in parts:
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            # Reference not found, return original
            return param

    # Return resolved parameter
    if isinstance(current, dict):
        return current

    return param


def format_parameter(
    param: Dict[str, Any],
    resolver: Optional[RefResolver] = None,
    swagger: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Format a parameter as a single-line bullet item.

    Format: {in}: {name} ({required/optional}) {type} [additional details]

    Examples:
    - query: queryLanguage (optional) string (enum: simple, full)
    - path: indexName (required) string
    - header: api-version (required) string
    - body: request (required) SearchRequest
    - ref: ClientRequestIdParameter (resolved or fallback)
    """
    # Resolve $ref if present
    if "$ref" in param:
        ref_name = param["$ref"].split("/")[-1]  # Extract name from path

        if swagger:
            resolved = resolve_parameter_ref(param, swagger)
            if resolved != param and "name" in resolved:
                # Successfully resolved, use resolved parameter
                param = resolved
            else:
                # Could not resolve, show ref name
                return f"ref: {ref_name} (optional)"
        else:
            # No swagger provided, show ref name
            return f"ref: {ref_name} (optional)"

    name = param.get("name", "")
    in_location = param.get("in", "")
    required = param.get("required", False)
    req_text = "required" if required else "optional"

    # Get type information
    type_info = ""
    if "schema" in param:
        type_info = format_type_info(param["schema"], resolver)
    elif "type" in param:
        # Swagger 2.0 style - type directly on parameter
        type_info = format_type_info(param, resolver)

    # Build base format
    parts = [f"{in_location}: {name} ({req_text})"]

    if type_info:
        parts.append(type_info)

    # Add enum info if present
    if "enum" in param and param.get("type") not in ["array", "object"]:
        enum_vals = param["enum"]
        if len(enum_vals) <= 3:
            parts.append(f"(enum: {', '.join(str(v) for v in enum_vals)})")

    return " ".join(parts)


def format_response(
    status_code: str, response: Dict[str, Any], resolver: Optional[RefResolver] = None
) -> str:
    """
    Format a response as a single-line bullet item.

    Format: {status_code}: {schema_name} ({content_type})

    Examples:
    - 200: SearchDocumentsResult (application/json)
    - 400: SearchError (application/json)
    - 201: (no content)
    """
    parts = [status_code]

    # Get schema information
    schema = response.get("schema")
    if schema:
        schema_name = format_type_info(schema, resolver)
        parts.append(schema_name)
    else:
        parts.append("(no content)")

    # Get content type from produces or default
    # Note: In Swagger 2.0, content types are typically in top-level "produces"
    # For simplicity, assume application/json
    parts.append("(application/json)")

    return ": ".join(parts[:2]) + " " + parts[2] if len(parts) > 2 else ": ".join(parts)


# ===== Operation Extraction =====


def extract_operations(
    swagger: Dict[str, Any], resolver: RefResolver
) -> Dict[str, Dict[str, Any]]:
    """
    Extract all operations from a swagger document.

    Returns a dictionary keyed by operation_key (path::method) with operation details:
    - path: API path
    - method: HTTP method
    - operationId: Operation identifier
    - parameters: List of parameter dicts
    - responses: Dict of status code -> response object
    - parameters_formatted: List of formatted parameter strings
    - responses_formatted: List of formatted response strings
    """
    operations = {}
    paths = swagger.get("paths", {})

    # Path-level parameters (inherited by all operations in the path)
    http_methods = {"get", "post", "put", "delete", "patch", "options", "head"}

    for path, path_obj in paths.items():
        # Get path-level parameters if any
        path_params = path_obj.get("parameters", [])

        for method, operation in path_obj.items():
            method_lower = method.lower()
            if method_lower not in http_methods:
                continue

            # Combine path-level and operation-level parameters
            op_params = operation.get("parameters", [])
            all_params = path_params + op_params

            # Format parameters (pass swagger for $ref resolution)
            params_formatted = []
            for param in all_params:
                params_formatted.append(format_parameter(param, resolver, swagger))

            # Format responses
            responses = operation.get("responses", {})
            responses_formatted = []
            for status_code, response in sorted(responses.items()):
                responses_formatted.append(
                    format_response(status_code, response, resolver)
                )

            # Create operation record
            op_key = get_operation_key(path, method_lower)
            operations[op_key] = {
                "path": path,
                "method": method_lower.upper(),
                "operationId": operation.get("operationId", ""),
                "parameters": all_params,
                "responses": responses,
                "parameters_formatted": params_formatted,
                "responses_formatted": responses_formatted,
            }

    return operations


# ===== Diff Computation =====


def compare_parameters(
    preview_params: List[Dict[str, Any]], ga_params: List[Dict[str, Any]]
) -> str:
    """
    Compare parameter lists between Preview and GA.

    Returns a summary string describing differences:
    - Added parameters: +paramName
    - Removed parameters: -paramName
    - Changed parameters: paramName (changed: detail)

    Examples:
    - "+speller, +semanticFields"
    - "-top"
    - "select (type changed)"
    - ""  (no changes)
    """
    # Build param dictionaries keyed by name::in
    preview_param_dict = {get_param_key(p): p for p in preview_params}
    ga_param_dict = {get_param_key(p): p for p in ga_params}

    preview_keys = set(preview_param_dict.keys())
    ga_keys = set(ga_param_dict.keys())

    changes = []

    # Added parameters (in Preview, not in GA)
    added = preview_keys - ga_keys
    for key in sorted(added):
        param = preview_param_dict[key]
        changes.append(f"+{param.get('name', key)}")

    # Removed parameters (in GA, not in Preview)
    removed = ga_keys - preview_keys
    for key in sorted(removed):
        param = ga_param_dict[key]
        changes.append(f"-{param.get('name', key)}")

    # Changed parameters (present in both)
    common = preview_keys & ga_keys
    for key in sorted(common):
        preview_param = preview_param_dict[key]
        ga_param = ga_param_dict[key]

        # Check for meaningful changes
        param_changes = []

        # Required change
        if preview_param.get("required") != ga_param.get("required"):
            old_req = "required" if ga_param.get("required") else "optional"
            new_req = "required" if preview_param.get("required") else "optional"
            param_changes.append(f"{old_req}→{new_req}")

        # Type change (simplified check)
        preview_type = preview_param.get("type") or preview_param.get("schema", {}).get(
            "type"
        )
        ga_type = ga_param.get("type") or ga_param.get("schema", {}).get("type")
        if preview_type != ga_type:
            param_changes.append(f"type: {ga_type}→{preview_type}")

        if param_changes:
            param_name = preview_param.get("name", key.split("::")[0])
            changes.append(f"{param_name} ({', '.join(param_changes)})")

    return ", ".join(changes) if changes else ""


def compare_responses(
    preview_responses: Dict[str, Any],
    ga_responses: Dict[str, Any],
    preview_resolver: Optional[RefResolver] = None,
    ga_resolver: Optional[RefResolver] = None,
) -> str:
    """
    Compare response definitions between Preview and GA.

    Returns a summary string describing differences:
    - Added status codes: +202
    - Removed status codes: -404
    - Changed response schemas: 200 (schema: A→B)

    Examples:
    - "+202"
    - "200 (schema: SearchResult→SearchDocumentsResult)"
    - ""  (no changes)
    """
    preview_codes = set(preview_responses.keys())
    ga_codes = set(ga_responses.keys())

    changes = []

    # Added response codes
    added = preview_codes - ga_codes
    for code in sorted(added):
        changes.append(f"+{code}")

    # Removed response codes
    removed = ga_codes - preview_codes
    for code in sorted(removed):
        changes.append(f"-{code}")

    # Changed response schemas
    common = preview_codes & ga_codes
    for code in sorted(common):
        preview_resp = preview_responses[code]
        ga_resp = ga_responses[code]

        # Compare schemas
        preview_schema = preview_resp.get("schema", {})
        ga_schema = ga_resp.get("schema", {})

        # Get schema identifiers for comparison
        preview_schema_id = format_type_info(preview_schema, preview_resolver)
        ga_schema_id = format_type_info(ga_schema, ga_resolver)

        if preview_schema_id != ga_schema_id and preview_schema_id and ga_schema_id:
            changes.append(f"{code} (schema: {ga_schema_id}→{preview_schema_id})")

    return ", ".join(changes) if changes else ""


def build_detail_diff(
    preview_op: Dict[str, Any],
    ga_op: Optional[Dict[str, Any]],
    preview_resolver: RefResolver,
    ga_resolver: Optional[RefResolver],
    preview_swagger: Dict[str, Any],
    ga_swagger: Optional[Dict[str, Any]],
) -> str:
    """
    Build a structured detail diff string with multiple sections.

    Returns a multi-line string with sections:
    - Path/Method/Operation
    - Parameters
    - Responses

    Example output:
    Path/Method/Operation
    • GA presence: missing
    • operationId: GA=n/a, Preview=Aliases_Create

    Parameters
    • + query: speller (optional) string
    • - header: x-ms-client-request-id (optional)

    Responses
    • + 202
    • ~ 200 schema changed: A -> B
    """
    sections = []

    # === Path/Method/Operation Section ===
    section_lines = ["Path/Method/Operation"]

    if ga_op:
        section_lines.append("• GA presence: present")

        # OperationId comparison
        ga_op_id = ga_op.get("operationId", "")
        preview_op_id = preview_op.get("operationId", "")
        if ga_op_id != preview_op_id:
            section_lines.append(
                f"• operationId: GA={ga_op_id or 'n/a'}, Preview={preview_op_id or 'n/a'}"
            )
        else:
            section_lines.append(f"• operationId: {preview_op_id}")
    else:
        section_lines.append("• GA presence: missing")
        preview_op_id = preview_op.get("operationId", "")
        section_lines.append(f"• operationId: GA=n/a, Preview={preview_op_id}")

    sections.append("\n".join(section_lines))

    # === Parameters Section ===
    if ga_op:
        # Build param dictionaries keyed by name::in
        preview_params = preview_op.get("parameters", [])
        ga_params = ga_op.get("parameters", [])

        # Resolve $refs for better comparison
        preview_params_resolved = [
            resolve_parameter_ref(p, preview_swagger) for p in preview_params
        ]
        ga_params_resolved = [
            resolve_parameter_ref(p, ga_swagger) if ga_swagger else p for p in ga_params
        ]

        preview_param_dict = {
            get_param_key(p): p for p in preview_params_resolved if "name" in p
        }
        ga_param_dict = {get_param_key(p): p for p in ga_params_resolved if "name" in p}

        preview_keys = set(preview_param_dict.keys())
        ga_keys = set(ga_param_dict.keys())

        param_lines = ["Parameters"]
        has_changes = False

        # Added parameters
        added = preview_keys - ga_keys
        for key in sorted(added):
            param = preview_param_dict[key]
            formatted = format_parameter(param, preview_resolver, preview_swagger)
            param_lines.append(f"• + {formatted}")
            has_changes = True

        # Removed parameters
        removed = ga_keys - preview_keys
        for key in sorted(removed):
            param = ga_param_dict[key]
            formatted = format_parameter(param, ga_resolver, ga_swagger)
            param_lines.append(f"• - {formatted}")
            has_changes = True

        # Changed parameters
        common = preview_keys & ga_keys
        for key in sorted(common):
            preview_param = preview_param_dict[key]
            ga_param = ga_param_dict[key]

            changes = []

            # Required change
            if preview_param.get("required") != ga_param.get("required"):
                old_req = "required" if ga_param.get("required") else "optional"
                new_req = "required" if preview_param.get("required") else "optional"
                changes.append(f"{old_req}→{new_req}")

            # Type change
            preview_type = preview_param.get("type") or preview_param.get(
                "schema", {}
            ).get("type")
            ga_type = ga_param.get("type") or ga_param.get("schema", {}).get("type")
            if preview_type != ga_type and preview_type and ga_type:
                changes.append(f"type {ga_type}→{preview_type}")

            if changes:
                param_name = preview_param.get("name", key.split("::")[0])
                param_in = preview_param.get(
                    "in", key.split("::")[1] if "::" in key else ""
                )
                change_text = ", ".join(changes)
                param_lines.append(f"• ~ {param_in}: {param_name} ({change_text})")
                has_changes = True

        if not has_changes:
            param_lines.append("• (no changes)")

        sections.append("\n".join(param_lines))
    else:
        # New operation - don't show parameter details
        sections.append("Parameters\n• (new operation)")

    # === Responses Section ===
    if ga_op:
        preview_responses = preview_op.get("responses", {})
        ga_responses = ga_op.get("responses", {})

        preview_codes = set(preview_responses.keys())
        ga_codes = set(ga_responses.keys())

        resp_lines = ["Responses"]
        has_changes = False

        # Added responses
        added = preview_codes - ga_codes
        for code in sorted(added):
            resp_lines.append(f"• + {code}")
            has_changes = True

        # Removed responses
        removed = ga_codes - preview_codes
        for code in sorted(removed):
            resp_lines.append(f"• - {code}")
            has_changes = True

        # Changed response schemas
        common = preview_codes & ga_codes
        for code in sorted(common):
            preview_resp = preview_responses[code]
            ga_resp = ga_responses[code]

            preview_schema = preview_resp.get("schema", {})
            ga_schema = ga_resp.get("schema", {})

            preview_schema_id = format_type_info(preview_schema, preview_resolver)
            ga_schema_id = format_type_info(ga_schema, ga_resolver)

            if preview_schema_id != ga_schema_id and preview_schema_id and ga_schema_id:
                resp_lines.append(
                    f"• ~ {code} schema changed: {ga_schema_id} → {preview_schema_id}"
                )
                has_changes = True

        if not has_changes:
            resp_lines.append("• (no changes)")

        sections.append("\n".join(resp_lines))
    else:
        # New operation
        sections.append("Responses\n• (new operation)")

    return "\n\n".join(sections)


def annotate_operations_with_ga_diff(
    preview_operations: Dict[str, Dict[str, Any]],
    ga_operations: Dict[str, Dict[str, Any]],
    preview_resolver: RefResolver,
    ga_resolver: RefResolver,
    preview_swagger: Dict[str, Any],
    ga_swagger: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """
    Annotate Preview operations with GA comparison data.

    Returns a list of operation records with diff annotations:
    - All fields from extract_operations()
    - ga_presence: "present" or "missing"
    - detail_diff: structured diff with Path/Method/Operation, Parameters, and Responses sections
    """
    annotated_ops = []

    for op_key, preview_op in sorted(preview_operations.items()):
        # Check if operation exists in GA
        ga_op = ga_operations.get(op_key)

        # Build detail diff
        detail_diff = build_detail_diff(
            preview_op,
            ga_op,
            preview_resolver,
            ga_resolver,
            preview_swagger,
            ga_swagger,
        )

        # Create annotated record with new GA presence labels
        annotated_op = {
            **preview_op,
            "ga_presence": "present in 2025-09-01"
            if ga_op
            else "new in 2025-11-01-preview",
            "detail_diff": detail_diff,
        }

        annotated_ops.append(annotated_op)

    return annotated_ops


# ===== Excel Generation =====


def create_excel_report(
    operations_data: Dict[str, List[Dict[str, Any]]], output_path: Path
) -> None:
    """
    Create Excel workbook with operation surface report.

    operations_data: Dict mapping file name -> list of annotated operations
    output_path: Path to output .xlsx file
    """
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Define column headers
    headers = [
        "2025-11-01-preview Path",
        "Method",
        "OperationId",
        "Parameters",
        "Responses",
        "2025-09-01 GA Presence",
        "Detail Diff",
        "2026-04-01 Decision",
        "TSP Action",
    ]

    # Header formatting
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data alignment (wrap text, top-aligned for multi-line cells)
    data_alignment = Alignment(vertical="top", wrap_text=True)

    # Color for missing operations
    missing_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    # Create a sheet for each file (in stable order: searchservice, searchindex, knowledgebase)
    sheet_order = ["searchservice.json", "searchindex.json", "knowledgebase.json"]
    for file_name in sheet_order:
        if file_name not in operations_data:
            continue

        operations = operations_data[file_name]
        # Sort operations by path then method for stable ordering
        operations = sorted(operations, key=lambda op: (op["path"], op["method"]))

        # Create sheet name (remove .json extension)
        sheet_name = file_name.replace(".json", "")
        ws = wb.create_sheet(title=sheet_name)

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_num, op in enumerate(operations, 2):
            # Basic operation info
            ws.cell(row=row_num, column=1, value=op["path"])
            ws.cell(row=row_num, column=2, value=op["method"])
            ws.cell(row=row_num, column=3, value=op["operationId"])

            # Parameters (bullet list, one per line)
            params_text = "\n".join(f"• {p}" for p in op["parameters_formatted"])
            params_cell = ws.cell(row=row_num, column=4, value=params_text)
            params_cell.alignment = data_alignment

            # Responses (bullet list, one per line)
            responses_text = "\n".join(f"• {r}" for r in op["responses_formatted"])
            responses_cell = ws.cell(row=row_num, column=5, value=responses_text)
            responses_cell.alignment = data_alignment

            # GA comparison annotations
            ws.cell(row=row_num, column=6, value=op["ga_presence"])

            # Detail Diff (multi-section structured diff)
            detail_diff_cell = ws.cell(row=row_num, column=7, value=op["detail_diff"])
            detail_diff_cell.alignment = data_alignment

            # Manual review columns (blank)
            ws.cell(row=row_num, column=8, value="")
            ws.cell(row=row_num, column=9, value="")

            # Highlight new operations (not in GA)
            if op["ga_presence"] == "new in 2025-11-01-preview":
                for col_num in range(1, 10):
                    ws.cell(row=row_num, column=col_num).fill = missing_fill

        # Auto-adjust column widths
        column_widths = {
            1: 40,  # 2025-11-01-preview Path
            2: 10,  # Method
            3: 35,  # OperationId
            4: 60,  # Parameters
            5: 50,  # Responses
            6: 20,  # 2025-09-01 GA Presence
            7: 70,  # Detail Diff (wider for structured content)
            8: 30,  # 2026-04-01 Decision
            9: 30,  # TSP Action
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Set row height for data rows (allow multi-line content)
        for row_num in range(2, len(operations) + 2):
            ws.row_dimensions[row_num].height = None  # Auto-height

    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel report saved: {output_path}")


# ===== Main Processing =====


def process_file_pair(
    preview_path: str, ga_path: Optional[str], file_name: str
) -> List[Dict[str, Any]]:
    """
    Process a single Preview/GA file pair and return annotated operations.

    If ga_path is None, all operations are marked as "missing" in GA.
    """
    print(f"\nProcessing {file_name}...")

    # Load Preview swagger
    preview_swagger = load_swagger_file(preview_path)
    preview_resolver = RefResolver(preview_swagger)

    # Extract Preview operations
    preview_operations = extract_operations(preview_swagger, preview_resolver)
    print(f"  Preview operations: {len(preview_operations)}")

    # Load GA swagger if provided
    if ga_path:
        try:
            ga_swagger = load_swagger_file(ga_path)
            ga_resolver = RefResolver(ga_swagger)
            ga_operations = extract_operations(ga_swagger, ga_resolver)
            print(f"  GA operations: {len(ga_operations)}")
        except FileNotFoundError:
            print(f"  Warning: GA file not found, treating as empty")
            ga_swagger = {}
            ga_operations = {}
            ga_resolver = RefResolver({})
    else:
        print(f"  No GA file provided, treating as empty")
        ga_swagger = {}
        ga_operations = {}
        ga_resolver = RefResolver({})

    # Annotate operations with GA diff (pass swagger documents for $ref resolution)
    annotated_operations = annotate_operations_with_ga_diff(
        preview_operations,
        ga_operations,
        preview_resolver,
        ga_resolver,
        preview_swagger,
        ga_swagger,
    )

    print(f"  Annotated operations: {len(annotated_operations)}")

    return annotated_operations


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Generate operation-centric API surface Excel report from Preview swagger with GA diff annotations"
    )

    # Note: --file argument is deprecated; the tool now always generates
    # a single workbook with 3 sheets (searchservice, searchindex, knowledgebase)
    # Kept for backward compatibility but ignored.
    parser.add_argument(
        "--file",
        choices=["searchindex", "searchservice", "knowledgebase", "all"],
        default="all",
        help="(Deprecated) Tool now always processes all files into one workbook",
    )

    # Custom file paths
    parser.add_argument("--preview-index", help="Path to preview searchindex.json")
    parser.add_argument("--ga-index", help="Path to GA searchindex.json")
    parser.add_argument("--preview-service", help="Path to preview searchservice.json")
    parser.add_argument("--ga-service", help="Path to GA searchservice.json")
    parser.add_argument(
        "--preview-knowledgebase", help="Path to preview knowledgebase.json"
    )
    parser.add_argument(
        "--ga-knowledgebase", help="Path to GA knowledgebase.json (optional)"
    )

    # Output options
    parser.add_argument(
        "--output-dir",
        default="./output",
        help="Output directory for Excel reports (default: ./output)",
    )

    args = parser.parse_args()

    # Default file paths
    script_dir = Path(__file__).parent
    default_preview_base = (
        script_dir
        / ".."
        / "data-plane"
        / "Search"
        / "preview"
        / "2025-11-01-preview_unmigrated"
    )
    default_ga_base = (
        script_dir / ".." / "data-plane" / "Search" / "stable" / "2025-09-01"
    )

    # Always process all 3 files to generate a single workbook with 3 sheets
    files_to_process = {}

    # searchservice
    preview_service = args.preview_service or str(
        default_preview_base / "searchservice.json"
    )
    ga_service = args.ga_service or str(default_ga_base / "searchservice.json")
    files_to_process["searchservice.json"] = (preview_service, ga_service)

    # searchindex
    preview_index = args.preview_index or str(default_preview_base / "searchindex.json")
    ga_index = args.ga_index or str(default_ga_base / "searchindex.json")
    files_to_process["searchindex.json"] = (preview_index, ga_index)

    # knowledgebase (GA doesn't exist, use None)
    preview_kb = args.preview_knowledgebase or str(
        default_preview_base / "knowledgebase.json"
    )
    ga_kb = args.ga_knowledgebase  # Usually None
    files_to_process["knowledgebase.json"] = (preview_kb, ga_kb)

    # Process files
    operations_data = {}

    for file_name, (preview_path, ga_path) in files_to_process.items():
        try:
            annotated_ops = process_file_pair(preview_path, ga_path, file_name)
            operations_data[file_name] = annotated_ops
        except FileNotFoundError as e:
            print(f"  Error: {e}")
            continue
        except Exception as e:
            print(f"  Error processing {file_name}: {e}")
            import traceback

            traceback.print_exc()
            continue

    if not operations_data:
        print("\nNo data to report. Exiting.")
        return

    # Generate Excel report
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"operation_surface_{timestamp}.xlsx"

    print(f"\nGenerating Excel report...")
    create_excel_report(operations_data, output_file)

    # Summary (display in sheet order)
    print("\n" + "=" * 60)
    print("Operation Surface Report Summary")
    print("=" * 60)
    sheet_order = ["searchservice.json", "searchindex.json", "knowledgebase.json"]
    for file_name in sheet_order:
        if file_name not in operations_data:
            continue
        operations = operations_data[file_name]
        present_count = sum(
            1 for op in operations if op["ga_presence"] == "present in 2025-09-01"
        )
        new_count = sum(
            1 for op in operations if op["ga_presence"] == "new in 2025-11-01-preview"
        )
        print(f"\n{file_name}:")
        print(f"  Total operations: {len(operations)}")
        print(f"  Present in 2025-09-01: {present_count}")
        print(f"  New in 2025-11-01-preview: {new_count}")

    print("\n" + "=" * 60)
    print(f"Report saved to: {output_file}")
    print("=" * 60)


if __name__ == "__main__":
    main()
